class NewUser:
    def dataInser():
        import tkinter as tk
        from openpyxl import Workbook, load_workbook
        from DataAdd import deposit as d

        # def encrypt():
        #     exit()

        def next1():
            a = submit()
            root.destroy()
            d.DataAddNewUser(a)

        def Home():
            wb.save('DataBase.xlsx')
            root.destroy()
            from WindowMain import main as m
            m.run()

        def submit():
            ws['G2'] = (ws['G2'].value) + 1
            Id = name_var.get()
            password = passw_var.get()
            # print(Id)
            num = ws['G2'].value
            ws[f'A{num}'] = Id
            ws[f'C{num}'] = password
            ws[f'E{num}'] = num
            Usernum = ws[f'E{num}'].value
            ws['V2'] = Usernum
            name_var.set("")
            passw_var.set("")
            wb.save('DataBase.xlsx')
            return Usernum

        root = tk.Tk()

        root.geometry("700x438")
        root.maxsize(700, 438)
        root.minsize(700, 438)
        root.title('Login Page')
        root.iconbitmap('Custom-Icon-Design-Pretty-Office-11-Coins.ico')
        root.config(bg='#B0CFDE')
        name_var = tk.StringVar()
        passw_var = tk.StringVar()

        wb = load_workbook('DataBase.xlsx')
        ws = wb['Login Info']
        position = 0
        Usernum = 0

        name_label = tk.Label(root, text='Username',
                              font=('Times New Roman', 10, 'bold'))
        name_entry = tk.Entry(root, textvariable=name_var,
                              font=('Times New Roman', 10, 'normal'))
        passw_label = tk.Label(root, text='Password',
                               font=('Times New Roman', 10, 'bold'))
        passw_entry = tk.Entry(root, textvariable=passw_var,
                               font=('Times New Roman', 10, 'normal'), show='*')
        sub_btn = tk.Button(root, text='Submit', command=submit)
        # ext_btn = tk.Button(root, text='Exit', command=encrypt)
        next_btn = tk.Button(root, text='Next', command=next1)
        home_btn = tk.Button(root,text='Home',command=Home)
        
        home_btn.place(x=600,y=400)
        name_label.place(x=50, y=150)
        name_entry.place(x=180, y=150)
        passw_label.place(x=50, y=200)
        passw_entry.place(x=180, y=200)
        next_btn.place(x=300, y=250)
        # ext_btn.place(x=400, y=250)
        sub_btn.place(x=500, y=250)

        root.mainloop()


class existingUser:

    def dataCheck():
        import tkinter as tk
        from openpyxl import Workbook, load_workbook
        from DataAdd import excistinUserDeposit as e
        from WithDrawal import caseWithDrawal as w

        root = tk.Tk()

        root.geometry("700x438")
        root.maxsize(700, 438)
        root.minsize(700, 438)
        root.title('Login Page')
        root.iconbitmap('Custom-Icon-Design-Pretty-Office-11-Coins.ico')
        root.config(bg='#B0CFDE')
        name_var = tk.StringVar()
        passw_var = tk.StringVar()

        wb = load_workbook('DataBase.xlsx')
        ws = wb['Login Info']
        position = 0
        Usernum = 0

        # def encrypt():
        #     exit()

        def next1():
            a = submit()
            root.destroy()
            e.DataAddExistingUser(a)

        def remove():
            a = submit()
            root.destroy()
            w.WithDrawalcase(a)

        def submit():
            Id = name_var.get()
            Input_password = passw_var.get()
            try:
                userId = Id
                i = 1
                for data in ws:
                    if ws[f'A{i}'].value == userId:
                        position = i
                        password = Input_password
                        if ws[f'C{position}'].value == password:
                            Usernum = position
                            ws['V2'] = Usernum
                            break

                        else:
                            my_label = tk.Label(root, text='Password Not Match !', fg='red', bg='#B0CFDE', font=(
                                'Times New Roman', 10, 'bold')).pack()
                    i += 1
                if(Usernum == None):
                    my_label2 = tk.Label(
                        root, text='Not Found ', fg='black').pack()
            except:
                # print("Something Went Wrong")
                my_label2 = tk.Label(root, text='Not Found !', fg='red', bg='#B0CFDE', font=(
                    'Times New Roman', 10, 'bold')).pack()
            name_var.set("")
            passw_var.set("")
            wb.save('DataBase.xlsx')
            return Usernum

        name_label = tk.Label(root, text='Username',
                              font=('Times New Roman', 10, 'bold'))
        name_entry = tk.Entry(root, textvariable=name_var,
                              font=('Times New Roman', 10, 'normal'))
        passw_label = tk.Label(root, text='Password',
                               font=('Times New Roman', 10, 'bold'))
        passw_entry = tk.Entry(root, textvariable=passw_var,
                               font=('Times New Roman', 10, 'normal'), show='*')
        # sub_btn = tk.Button(root, text='Submit', command=submit)
        # ext_btn = tk.Button(root, text='Exit', command=encrypt)
        next_btn = tk.Button(root, text='Deposit', command=next1)
        remove_btn = tk.Button(root, text='Withdrawal', command=remove)

        name_label.place(x=50, y=150)
        name_entry.place(x=180, y=150)
        passw_label.place(x=50, y=200)
        passw_entry.place(x=180, y=200)
        # sub_btn.place(x=300, y=250)
        # ext_btn.place(x=500, y=250)
        next_btn.place(x=300, y=250)
        remove_btn.place(x=390, y=250)

        root.mainloop()

# dataCheck()
