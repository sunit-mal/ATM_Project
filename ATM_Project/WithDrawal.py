class caseWithDrawal:
    def WithDrawalcase(Usernum):
            from openpyxl import Workbook, load_workbook
            import tkinter as tk
            

            # def encrypt():
            #     exit()

            def submit():
                amountstr = amount_var.get()
                amount = int(amountstr)
                ws[f'F{Usernum}']=str(ws[f'F{Usernum}'].value)+';'+str(ws[f'E{Usernum}'].value)+'-'+str(amount)
                ws[f'E{Usernum}'] = int(ws[f'E{Usernum}'].value) - amount
                wb.save('DataBase.xlsx')
                amount_var.set('')
            def show():
                Balance = ws[f'E{Usernum}'].value
                my_lebel = tk.Label(root,text=Balance,font=('Times New Roman',10,'bold'))
                my_lebel.place(x=300,y=350)

            def Home():
                wb.save('DataBase.xlsx')
                root.destroy()
                from WindowMain import main as m
                m.run()

            wb = load_workbook('DataBase.xlsx')
            ws = wb['Balance Sheet']
            

            if(ws[f'A{Usernum}'].value) == None:
                deposit.DataAddNewUser(Usernum)
            root = tk.Tk()

            root.geometry("700x438")
            root.maxsize(700, 438)
            root.minsize(700, 438)
            root.title('Withdrawal Window')
            root.iconbitmap('Custom-Icon-Design-Pretty-Office-11-Coins.ico')
            root.config(bg='#B0CFDE')
            amount_var = tk.StringVar()

            amount_label = tk.Label(root, text='Enter Amount', font=('Times New Roman', 10, 'bold'))
            amount_entry = tk.Entry(root, textvariable=amount_var,
                                    font=('Times New Roman', 10, 'normal'))
            sub_btn = tk.Button(root, text='Submit', command=submit)
            # ext_btn = tk.Button(root, text='Exit', command=encrypt)
            show_btn = tk.Button(root,text='Show Balance',command=show)
            home_btn = tk.Button(root,text='Home',command=Home)
                
            amount_label.place(x=50, y=200)
            amount_entry.place(x=180, y=200)
            sub_btn.place(x=300, y=250)
            # ext_btn.place(x=400, y=250)
            show_btn.place(x=400,y=350)
            home_btn.place(x=600,y=400)
            root.mainloop()


# DataAddExistingUser(2)