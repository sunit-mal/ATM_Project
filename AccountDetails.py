#   Window - 4
class BioData:
    def Run(Usernum):
        import tkinter as tkk
        from openpyxl import Workbook, load_workbook
        from PIL import Image,ImageTk

        def nextbtn():
            name=name_var.get()
            Mobile=Mobile_var.get()
            email=email_var.get()

            ws[f'A{Usernum}'] = name
            ws[f'B{Usernum}'] =Mobile
            ws[f'D{Usernum}'] =email
            wb.save('DataBase.xlsx')
            root.destroy()
            from Deposit import DepositNewUser as data
            data.Run(Usernum)

        def clean():
            name_var.set('')
            Mobile_var.set('')
            email_var.set('')

        wb = load_workbook('DataBase.xlsx')
        ws = wb['CustomerData']
        root=tkk.Tk()
        root.title("Bio Data")
        root.geometry("1280x720")
        root.configure(bg='lightblue')
        root.resizable(False, False)
        root.iconbitmap('./download.ico')
        img= Image.open('Background.jpg')
        img_file= img.resize((1280,720), Image.ANTIALIAS)
        bg = ImageTk.PhotoImage(img_file)
        bgl = tkk.Label(root,image=bg,bd=0)
        bgl.place(x=0, y=0, width=1280,height=720)

        name_var = tkk.StringVar()
        Mobile_var = tkk.StringVar()
        email_var = tkk.StringVar()

        tkk.Label(root,text="Name:", font=("Times New Roman", 40)).place(x=200,y=150)
        tkk.Label(root,text="Mobile No:", font=("Times New Roman", 40)).place(x=200,y=255)
        tkk.Label(root,text="Email:", font=("Times New Roman", 40)).place(x=200,y=360)
        tkk.Entry(root,textvariable=name_var,bd=4, font=("Times New Roman", 30)).place(x=600,y=150)
        tkk.Entry(root,textvariable=Mobile_var,bd=4,font=("Times New Roman", 30)).place(x=600,y=255)
        tkk.Entry(root,textvariable=email_var,bd=4, font=("Times New Roman", 30)).place(x=600,y=360)
        tkk.Button(root,text="Clear", height=3, width=25, bd=5, font=("Helvetica"),command=clean).place(x=650,y=560)
        tkk.Button(root,text="Next",height=3,width=25,bd=5,font=("Helvetica"),command=nextbtn).place(x=950,y=560)
        wb.save('DataBase.xlsx')
        tkk.mainloop()
# BioData.Run(2)
