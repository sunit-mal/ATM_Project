#   Window - 11

class balanceenq:
    def Run(Usernum):
        from openpyxl import Workbook,load_workbook
        import tkinter as tk
        from PIL import Image,ImageTk
        def backHome():
                root.destroy()
                from SignType import User as H
                H.Run()

        wb = load_workbook('DataBase.xlsx')
        ws = wb['BalanceSheet']
        root = tk.Tk()
        root.title("Balance")
        root.geometry("1280x720")
        root.configure(bg='lightblue')
        root.resizable(False, False)
        root.iconbitmap('./download.ico')
        img= Image.open('Background.jpg')
        img_file= img.resize((1280,720), Image.ANTIALIAS)
        bg = ImageTk.PhotoImage(img_file)
        bgl = tk.Label(root,image=bg,bd=0)
        bgl.place(x=0, y=0, width=1280,height=720)

        Balance = ws[f'C{Usernum}'].value    #   Add Balance From DataBase 
        tk.Label(root, text="YOUR BALANCE IS:", fg='blue',font=("Times New Roman", 50)).place(x=380, y=270)
        tk.Label(root, text=Balance, fg='blue',font=("Times New Roman", 50)).place(x=600, y=370)
        tk.Button(root, text="OK", height=3, width=20, bd=5, font=("Times New Roman", 15, "bold"),command=backHome).place(x=1000, y=590)
        wb.save('DataBase.xlsx')
        tk.mainloop()

# balanceenq.Run(2)
