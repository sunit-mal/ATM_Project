#   Window - 6,10

class DepositNewUser:
    def Run(Usernum):
        import tkinter as tkk
        from openpyxl import Workbook, load_workbook

        def clean():
            amount_var.set('')

        def Home():
            root.destroy()
            from SignType import User as H
            H.Run()

        def submit():
            amountstr = amount_var.get()
            amount = int(amountstr)
            if (ws[f'C{Usernum}'].value == None):
                ws[f'A{Usernum}'] = ws1[f'A{Usernum}'].value
                ws[f'C{Usernum}'] = 0
            ws[f'F{Usernum}']=str(ws[f'F{Usernum}'].value)+';'+str(ws[f'C{Usernum}'].value)+'+'+str(amount)
            ws[f'C{Usernum}'] = int(ws[f'C{Usernum}'].value) + amount
            amount_var.set('')
            wb.save('DataBase.xlsx')
            root.destroy()
            from ThakyouDepositWindow import ThankYouWWindows as t
            t.play()

        wb = load_workbook('DataBase.xlsx')
        ws = wb['BalanceSheet']
        ws1 = wb['CustomerData']
        root=tkk.Tk()
        root.title("Deposit")
        root.geometry("1280x720")
        root.configure(bg='lightblue')
        root.resizable(False, False)
        root.iconbitmap('./download.ico')
        amount_var = tkk.StringVar()

        tkk.Label(root,text="Amount:", font=("Times New Roman", 40)).place(x=200,y=153)
        
        tkk.Entry(root,textvariable=amount_var,bd=4, font=("Times New Roman", 30)).place(x=600,y=153)

        tkk.Button(root,text="Home", height=3, width=25, bd=5, font=("Helvetica"),command=Home).place(x=145,y=560)
        tkk.Button(root,text="Clear", height=3, width=25, bd=5, font=("Helvetica"),command=clean).place(x=650,y=560)
        tkk.Button(root,text="Next",height=3,width=25,bd=5,font=("Helvetica"),command=submit).place(x=950,y=560)

        tkk.mainloop()

DepositNewUser.Run(3)