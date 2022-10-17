# Window 2 & 3

# Window - 2
class LoginPageForExsistingUser:
    def Run():
        import tkinter as tk
        from openpyxl import Workbook, load_workbook
        from PIL import Image,ImageTk

        wb = load_workbook('DataBase.xlsx')
        ws = wb['LoginInfo']
        def TransationNext():
            Id = UserId.get()
            Input_password = Password.get()
            try:
                userId = Id
                i = 1
                for data in ws:
                    if ws[f'A{i}'].value == userId:
                        position = i
                        password = Input_password
                        if ws[f'B{position}'].value == password:
                            Usernum = ws[f'C{position}'].value
                            root.destroy()
                            from Transationmenu import Transaction_Menu as l
                            l.Run(Usernum)
                            break

                        else:
                            my_label = tk.Label(root, text='Password Not Match !', fg='red', bg='#B0CFDE', font=(
                                'Times New Roman', 13, 'bold')).place(x=350,y=30)
                    i += 1
                if(ws[f'A{i}'].value == None):
                    my_label2 = tk.Label(
                        root, text='Not Found ',fg='red', bg='#B0CFDE', font=(
                    'Times New Roman', 13, 'bold')).place(x=350,y=10)
            except:
                my_label2 = tk.Label(root, text='Not Found !', fg='red', bg='#B0CFDE', font=(
                    'Times New Roman', 13, 'bold')).place(x=350,y=10)

        def welcome():
            root.destroy()
            from SignType import User as H
            H.Run()

        root=tk.Tk()
        root.title("Login")
        root.geometry("1280x720")
        root.configure(bg='lightblue')
        root.resizable(False, False)
        root.iconbitmap('./download.ico')
        img= Image.open('Background.jpg')
        img_file= img.resize((1280,720), Image.ANTIALIAS)
        bg = ImageTk.PhotoImage(img_file)
        bgl = tk.Label(root,image=bg,bd=0)
        bgl.place(x=0, y=0, width=1280,height=720)
        UserId = tk.StringVar()
        Password = tk.StringVar()

        tk.Label(root,text="User Id:", font=("Times New Roman", 30)).place(x=200,y=80)
        tk.Label(root,text="Password:", font=("Times New Roman", 30)).place(x=200,y=240)
        tk.Entry(root,bd=4,textvariable= UserId,width=30, font=("Times New Roman", 30)).place(x=600,y=85)
        tk.Entry(root,bd=4,textvariable= Password,width=30, show='*',font=("Times New Roman", 30)).place(x=600,y=245)
        tk.Button(root,text="Next", height=2, width=20, bd=5, font=("Times New Roman", 15,"bold"),command=TransationNext).place(x=650,y=560)
        tk.Button(root,text="Back",height=2,width=20,bd=5,font=("Times New Roman", 15,"bold"),command=welcome).place(x=950,y=560)
        tk.mainloop()

# Window - 3
class LoginPageForNewUser:
    def Run():
        import tkinter as tkk
        from openpyxl import Workbook, load_workbook
        from PIL import Image,ImageTk

        wb = load_workbook('DataBase.xlsx')
        ws = wb['LoginInfo']
        def SigninPage():
            root.destroy
            from SignType import User as n
            n.Run()
            
        def NewaccountNext():
            ws['H2'] = (ws['H2'].value) + 1
            Id = UserId.get()
            password = Password.get()
            # print(Id)
            num = ws['H2'].value
            ws[f'A{num}'] = Id
            ws[f'B{num}'] = password
            ws[f'C{num}'] = num
            Usernum = ws[f'C{num}'].value
            # ws[''] = Usernum
            UserId.set("")
            Password.set("")
            wb.save('DataBase.xlsx')
            root.destroy()
            from AccountDetails import BioData as y
            y.Run(Usernum)
            
        root=tkk.Tk()
        root.title("Login")
        root.geometry("1280x720")
        root.configure(bg='lightblue')
        root.resizable(False, False)
        root.iconbitmap('./download.ico')
        img= Image.open('Background.jpg')
        img_file= img.resize((1280,720), Image.ANTIALIAS)
        bg = ImageTk.PhotoImage(img_file)
        bgl = tkk.Label(root,image=bg,bd=0)
        bgl.place(x=0, y=0, width=1280,height=720)
        UserId = tkk.StringVar()
        Password = tkk.StringVar()

        tkk.Label(root,text="User Id:", font=("Times New Roman", 30)).place(x=200,y=80)
        tkk.Label(root,text="Password:", font=("Times New Roman", 30)).place(x=200,y=240)
        tkk.Entry(root,bd=7,textvariable= UserId,width=30, show='*',font=("Times New Roman", 30)).place(x=600,y=85)
        tkk.Entry(root,bd=7,textvariable= Password,width=30, show='*',font=("Times New Roman", 30)).place(x=600,y=245)
        tkk.Button(root,text="Next", height=2, width=20, bd=5, font=("Times New Roman", 15,"bold"),command=NewaccountNext).place(x=650,y=560)
        tkk.Button(root,text="Back",height=2,width=20,bd=5,font=("Times New Roman", 15,"bold"),command=SigninPage).place(x=950,y=560)
        wb.save('DataBase.xlsx')
        tkk.mainloop()

# LoginPageForExsistingUser.Run()
# LoginPageForNewUser.Run()