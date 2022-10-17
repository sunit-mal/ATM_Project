class VarificationCheck:
    def send(num,passw):
        from twilio.rest import Client
        import random
        from openpyxl import Workbook, load_workbook
        from VarificationWindow import OTPGENERATOR as var

        wb = load_workbook('DataBase.xlsx')
        ws = wb['CustomerData']
        number = str(ws[f'B{num}'].value)

        account_sid = 'ACf28401468eb08992f2e5bce883d5351a'
        auth_token='048fefbd35f286ee17453d5c346cb88c'
        twilio_number='+16076693466'
        target_number='+91'+number


        otp = random.randint(1000,9999)

        client = Client(account_sid,auth_token)

        massage = client.messages.create(
            body="Your Varification Code is: "+str(otp),
            from_=twilio_number,
            to=target_number
            )
        
        wb.save('DataBase.xlsx')
        var.Run(otp,passw,num) 
# VarificationCheck.send(3, 1122)