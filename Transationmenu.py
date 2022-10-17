#   Window - 5

class Transaction_Menu:
    def Run(Usernum):
        import tkinter as tkk
        from openpyxl import Workbook, load_workbook
        from PIL import Image,ImageTk
        
        wb = load_workbook('DataBase.xlsx')
        ws = wb['LoginInfo']
        Usernum = ws[f'C{Usernum}'].value
        def Withdrawl():
            root.destroy()
            from WithdrawlPage import Withdrawl as w
            w.Run(Usernum)
        def PinChange():
            root.destroy()
            from PinChange import  pin_chnage as w
            w.Run(Usernum)

        def BalanceEnq():
            root.destroy()
            from BalanceEnq import balanceenq as w
            w.Run(Usernum)
        def Deposit():
            root.destroy()
            from Deposit import DepositNewUser as w 
            w.Run(Usernum) 
        def Home():
            root.destroy()
            from SignType import User as H
            H.Run()        

        root=tkk.Tk()
        root.title("Transaction Menu")
        root.geometry("1280x720")
        root.configure(bg='lightblue')
        root.resizable(False, False)
        root.iconbitmap('./download.ico')
        img= Image.open('Background.jpg')
        img_file= img.resize((1280,720), Image.ANTIALIAS)
        bg = ImageTk.PhotoImage(img_file)
        bgl = tkk.Label(root,image=bg,bd=0)
        bgl.place(x=0, y=0, width=1280,height=720)

        tkk.Button(root,text="Balance Enquiry", height=3, width=40, bd=5, font=("Helvetica"),command=BalanceEnq).place(x=50,y=100)
        tkk.Button(root,text="Deposit",height=3,width=40,bd=5,font=("Helvetica"),command=Deposit).place(x=50,y=400)
        tkk.Button(root,text="Cash Withdrawl", height=3, width=40, bd=5, font=("Helvetica") ,command=Withdrawl).place(x=850,y=100)
        tkk.Button(root,text="Pin Change",height=3,width=40,bd=5,font=("Helvetica"),command=PinChange).place(x=850,y=400)
        tkk.Button(root,text="Home",height=3,width=25,bd=5,font=("Helvetica"),command=Home).place(x=515,y=600)
        wb.save('DataBase.xlsx')
        tkk.mainloop()

# Transaction_Menu.Run(3)