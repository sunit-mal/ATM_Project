class OTPGENERATOR:
    def Run(num,passw,Usernum):
        import tkinter as tk
        from openpyxl import Workbook, load_workbook
        from PIL import Image,ImageTk

        def verifing():
            # print(num,passw)
            otp = OTPNumber.get()
            if (otp == str(num)):
                ws[f'B{Usernum}'] = str(passw)
                wb.save('DataBase.xlsx')
                
                root.destroy()
                from pinchangedone import pin_change_done as sh
                sh.play()
            else:
                OTPNumber.set('')

        wb = load_workbook('DataBase.xlsx')
        ws = wb['LoginInfo']
        root=tk.Tk()
        root.title("OTPGENERATOR")
        root.geometry("1280x720")
        root.configure(bg='lightblue')
        root.resizable(False, False)
        root.iconbitmap('./download.ico')
        img= Image.open('Background.jpg')
        img_file= img.resize((1280,720), Image.ANTIALIAS)
        bg = ImageTk.PhotoImage(img_file)
        bgl = tk.Label(root,image=bg,bd=0)
        bgl.place(x=0, y=0, width=1280,height=720)

        OTPNumber = tk.StringVar()

        tk.Label(root,text="ENTER OTP ", font=("Times New Roman", 30)).place(x=200,y=260)
        tk.Entry(root,bd=4,textvariable= OTPNumber,width=20, font=("Times New Roman", 30)).place(x=600,y=270) 
        tk.Button(root,text="Done", height=2, width=20, bd=5, font=("Times New Roman", 15,"bold"),command=verifing).place(x=650,y=460)       
        wb.save('DataBase.xlsx')
        tk.mainloop()

# OTPGENERATOR.Run(2121,'2002',2)


