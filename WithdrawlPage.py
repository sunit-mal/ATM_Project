#   Window - 7

class Withdrawl:
    def Run(Usernum):
        import tkinter as tkk
        from openpyxl import Workbook, load_workbook

        def submit():
                amountstr = amount_var.get()
                amount = int(amountstr)
                ws[f'F{Usernum}']=str(ws[f'F{Usernum}'].value)+';'+str(ws[f'C{Usernum}'].value)+'-'+str(amount)
                ws[f'C{Usernum}'] = int(ws[f'C{Usernum}'].value) - amount
                wb.save('DataBase.xlsx')
                # amount_var.set('')
                root.destroy()
                from collectmoney import collectmoney as c
                c.play()

        def clean():
            amount_var.set('')

        def backHome():
            root.destroy()
            from SignType import User as H
            H.Run()

        wb = load_workbook('DataBase.xlsx')
        ws = wb['BalanceSheet']   

        root=tkk.Tk()
        root.title("Withdrawl")
        root.geometry("1280x720")
        root.configure(bg='lightblue')
        root.resizable(False, False)
        root.iconbitmap('./download.ico')
        amount_var = tkk.StringVar()

        tkk.Label(root,text="Amount:", font=("Times New Roman", 40)).place(x=200,y=153)
        
        tkk.Entry(root,textvariable=amount_var,bd=4, font=("Times New Roman", 30)).place(x=600,y=153)

        tkk.Button(root,text="Home", height=3, width=25, bd=5, font=("Helvetica"),command=backHome).place(x=145,y=560)
        tkk.Button(root,text="Clear", height=3, width=25, bd=5, font=("Helvetica"),command=clean).place(x=650,y=560)
        tkk.Button(root,text="Next",height=3,width=25,bd=5,font=("Helvetica"),command=submit).place(x=950,y=560)

        tkk.mainloop()

Withdrawl.Run(2)
